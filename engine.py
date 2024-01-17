import copy
import os

import openpyxl
import xml.etree.ElementTree as eTree
import subprocess

from PyQt5.QtWidgets import QMessageBox


# Start full process of making card from xlsx and svg template.
def process_starter(svg, xlsx, files_name, files_path, files_format, inkscape_dir, replace_file,  dpi=500):

    try:
        workbook = openpyxl.load_workbook(xlsx)
        sheet = workbook.active
    except FileNotFoundError:
        QMessageBox.warning(None, "File not found", "Can't find xlsx file. Please check path.", QMessageBox.Ok)
        return False
    try:
        tree = eTree.parse(svg)
        root = tree.getroot()
    except FileNotFoundError:
        QMessageBox.warning(None, "File not found", "Can't find svg file. Please check path.", QMessageBox.Ok)
        return False

    # Started from 3 because numbers (program) export xlsx with table name as row and second row is for column name.
    # Iterate over rows each loop give another row with data to process.
    for i in range(3, sheet.max_row + 1):
        error = True
        for quantity in range(quantity_print(sheet, i)):
            new_root = root_modifier(sheet, root, i)
            if "%VAR_" in files_name:
                unique_name = name_receiver(sheet, i, files_name)
            else:
                unique_name = files_name

            if unique_name is None:
                error = False
            else:
                if files_format == "svg":
                    if svg_maker(files_path, unique_name, new_root, replace_file) is None:
                        error = False
                elif files_format == "png":
                    if png_maker(files_path, unique_name, new_root, dpi, inkscape_dir, replace_file) is None:
                        error = False
                elif files_format == "pdf":
                    if pdf_maker(files_path, unique_name, new_root, inkscape_dir, replace_file) is None:
                        error = False
        if not error:
            print("error")

def quantity_print(sheet, row_number):
    # Loop over rows that have name with data to search.
    for row in sheet.iter_rows(min_row=row_number, max_row=row_number):
        # Loop over columns (assuming the header is in the second row)
        for cell, name_cell in zip(row, sheet[2]):
            if "Quantity" in name_cell.value:
                if cell.value is None:
                    return 1
                else:
                    return int(cell.value)

    # If "Quantity" is not found in any of the cells, return a default value (e.g., 1)
    return 1


# Take xlsx data file as sheet, make copy of provided svg root and take row_number to iterate over specific row in xlsx.
# Return modified root.
def root_modifier(sheet, root, row_number):
    new_root = copy.deepcopy(root)
    for row in sheet.iter_rows(min_row=row_number, max_row=row_number):
        # Loop over row that have name with data to search.
        for column in sheet.iter_rows(min_row=2, max_row=2):
            # Take column name and current row value and save it then send to replace.
            for cell, name in zip(row, column):
                cell_value = str(cell.value)
                # Each text in svg should start with "%VAR_" and with "%".
                text_replace = "%VAR_" + str(name.value) + "%"
                picture_replace = str(name.value) + ".png"
                # Replace values method.
                replace_values(new_root, text_replace, picture_replace, cell_value)
    return new_root


# Replace values in xml_root with values from xlsx.
def replace_values(xml_root, name_to_change, new_picture_name, cell_value):
    # Inkscape delete from picture path some information that need to be recovered manually.
    # IF YOU(I) CHANGE FOLDER WHERE PICTURE ARE YOU NEED FIND WHAT INKSCAPE DELETE AND CHANGE THAT.
    namespace = {'sodipodi': 'http://sodipodi.sourceforge.net/DTD/sodipodi-0.dtd'}
    attribute = './/*[@sodipodi:absref]'
# Change text %VAR_some_text% to normal text in xlsx.
    for element in xml_root.iter():
        if element.text == name_to_change in element.text:
            element.text = element.text.replace(name_to_change, cell_value)
        # Search for image, take old path and make new with new file name from new_picture_name.
    for image_elem in xml_root.findall(attribute, namespace):
        old_full_path = image_elem.get("{http://sodipodi.sourceforge.net/DTD/sodipodi-0.dtd}absref")
        folder_path = os.path.dirname(old_full_path)
        new_image_path = os.path.join(folder_path, new_picture_name)
        if old_full_path == new_image_path:
            cleared_path = new_image_path.replace(new_picture_name, '')
            path_with_file = os.path.join(str(cleared_path), str(cell_value) + ".png")
            image_elem.set('{http://www.w3.org/1999/xlink}href', path_with_file)


# Return unique card name from column named "Name".
def name_receiver(sheet, row_number, column_name):
    for row in sheet.iter_rows(min_row=row_number, max_row=row_number):
        # Loop over row that have name with data to search.
        for column in sheet.iter_rows(min_row=2, max_row=2):
            # Take column name and current row value and save it.
            for cell, name in zip(row, column):
                if column_name == "%VAR_" + str(name.value) + "%":
                    if cell.value is None:
                        return None
                    card_name = str(cell.value)
                    return card_name


# Check if chosen file name exists if not add number incremented for each new copy.
# Return name that doesn't exist.
def valid_name(directory, file_name, extension, replace):
    original_name = file_name

    # If user has chosen to replace the file, return the original name.
    if replace and os.path.isfile(os.path.join(directory, file_name + extension)):
        return file_name + extension

    increment = 1
    existing_files = set()

    # Iterate through files in the directory and store existing names in a set.
    for root, dirs, files in os.walk(directory):
        for existing_file in files:
            existing_files.add(os.path.splitext(existing_file)[0])

    # Check if the base name already exists, and if so, increment the file_name.
    while file_name in existing_files:
        file_name = original_name + str(increment)
        increment += 1

    # Return the new file name.
    return file_name + extension


# Take xml tree, name and save new file svg to provided path.
def svg_maker(file_path, file_name, xml_root, rep_value):
    extension = ".svg"
    # Check names if exists add number before.
    file_name = valid_name(file_path, file_name, extension, rep_value)

    if file_name is None:
        return

    file_path = os.path.join(file_path, file_name)

    xml_string = eTree.tostring(xml_root, encoding="utf-8", method="xml").decode()

    try:
        # Save the SVG string to the specified file.
        with open(file_path, 'w') as svg_file:
            svg_file.write(xml_string)
    except FileNotFoundError:
        QMessageBox.warning(None, "File not found", "Can't find Directory. Please check path.", QMessageBox.Ok)
        return None

    return "Done"


# Create png based on temporary SVG file from xml_root, change dpi for better resolution.
def png_maker(file_path, file_name, xml_root, dpi, inkscape_exe, rep_value):
    extension = ".png"
    # Check names if exists add number before.
    file_name = valid_name(file_path, file_name, extension, rep_value)

    if file_name is None:
        return

    png_file_path = os.path.join(file_path, file_name)

    try:
        # Create temporary SVG file.
        svg_file_path = os.path.join(file_path, "temp.svg")
        with open(svg_file_path, 'w') as svg_file:
            svg_file.write(eTree.tostring(xml_root).decode())
    except FileNotFoundError:
        QMessageBox.warning(None, "File not found", "Can't find Directory. Please check path.", QMessageBox.Ok)
        os.remove(svg_file_path)
        return None

    try:
        # Create PNG file using Inkscape
        inkscape_cmd = [
            inkscape_exe,
            '--export-type=png',
            '--export-filename={}'.format(png_file_path),
            '--export-dpi={}'.format(dpi),
            svg_file_path
        ]
        # Run Inkscape.
        subprocess.run(inkscape_cmd)
    except FileNotFoundError:
        QMessageBox.warning(None, "File not found", "Can't find inkscape.exe. Please check path.", QMessageBox.Ok)
        os.remove(svg_file_path)
        return None
    except Exception as e:
        # Handle other exceptions (e.g., Inkscape execution failure)
        QMessageBox.warning(None, "Error", f"An error occurred: {e}", QMessageBox.Ok)
        os.remove(svg_file_path)
        return None

    # Remove the temporary SVG file.
    os.remove(svg_file_path)
    return png_file_path


# Create pdf based on temporary SVG file from xml_root.
def pdf_maker(file_path, file_name, xml_root, inkscape_exe, rep_value):
    extension = ".pdf"
    # Check names if exists add number before.
    file_name = valid_name(file_path, file_name, extension, rep_value)

    if file_name is None:
        return

    pdf_file_path = os.path.join(file_path, file_name)

    try:
        # Create temporary SVG file.
        svg_file_path = os.path.join(file_path, "temp.svg")
        with open(svg_file_path, 'w') as svg_file:
            svg_file.write(eTree.tostring(xml_root).decode())
    except FileNotFoundError:
        QMessageBox.warning(None, "File not found", "Can't find inkscape.exe. Please check path.", QMessageBox.Ok)
        os.remove(svg_file_path)
        return None

    try:
        # Create PDF file using Inkscape.
        inkscape_cmd = [
            inkscape_exe,
            '--export-type=pdf',
            '--export-filename={}'.format(pdf_file_path),
            svg_file_path
        ]
        # Run Inkscape.
        subprocess.run(inkscape_cmd)
    except FileNotFoundError:
        QMessageBox.warning(None, "File not found", "Can't find inkscape exe. Please check path.", QMessageBox.Ok)
        os.remove(svg_file_path)
        return None

    # Remove the temporary SVG file.
    os.remove(svg_file_path)
    return "Done"
